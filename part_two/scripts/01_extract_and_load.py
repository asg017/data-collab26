"""
01_extract_and_load.py
Extract Illinois and California education data, normalize, and load into DuckDB.

Illinois: Report Card Public Data Set 2024 (Excel, 949+ columns)
  - We extract chronic absenteeism rates + enrollment + school identifiers from the 'General' sheet.
  - Illinois stores data in WIDE format: one row per school/district, columns for each demographic breakdown.

California: Chronic Absenteeism 2023-24 (TSV text file)
  - LONG format: one row per school × reporting category (demographic group).
  - Columns: enrollment, count, rate for chronic absenteeism.

This script:
  1. Parses both files
  2. Normalizes Illinois wide-format into long-format to match California
  3. Loads both into a DuckDB database
"""

import duckdb
import pandas as pd
import openpyxl
import os
import sys

DATA_DIR = os.path.join(os.path.dirname(__file__), '..', 'data')
DB_PATH = os.path.join(DATA_DIR, 'education.duckdb')

# ---------------------------------------------------------------------------
# 1. Parse California data
# ---------------------------------------------------------------------------
def load_california():
    print("Loading California chronic absenteeism data...")
    ca = pd.read_csv(
        os.path.join(DATA_DIR, 'ca_chronic_absenteeism_2024.txt'),
        sep='\t',
        dtype=str,
        encoding='latin-1'
    )
    ca.columns = ca.columns.str.strip()
    print(f"  Rows: {len(ca):,}, Columns: {len(ca.columns)}")
    print(f"  Columns: {list(ca.columns)}")

    # Clean numeric columns
    for col in ['ChronicAbsenteeismEligibleCumulativeEnrollment', 'ChronicAbsenteeismCount', 'ChronicAbsenteeismRate']:
        ca[col] = pd.to_numeric(ca[col].replace('*', pd.NA), errors='coerce')

    # Add state identifier
    ca['State'] = 'CA'
    return ca


# ---------------------------------------------------------------------------
# 2. Parse Illinois data (wide -> long for chronic absenteeism)
# ---------------------------------------------------------------------------
def load_illinois():
    print("Loading Illinois Report Card data...")
    wb = openpyxl.load_workbook(os.path.join(DATA_DIR, 'il_report_card_2024.xlsx'), read_only=True)
    ws = wb['General']

    # Read headers
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]

    # Identify columns we need
    # Core identity columns (indices 0-8)
    id_cols = {
        0: 'RCDTS',
        1: 'Type',
        2: 'School Name',
        3: 'District',
        4: 'City',
        5: 'County',
        6: 'District Type',
        7: 'School Type',
        8: 'Grades Served',
        14: '# Student Enrollment',
    }

    # Find chronic absenteeism columns
    ca_cols = {}
    enrollment_cols = {}
    for i, h in enumerate(headers):
        if h and 'Chronic Absenteeism' in str(h) and i < 536:  # Skip CRDC columns
            ca_cols[i] = str(h)
        if h and '# Student Enrollment' in str(h):
            enrollment_cols[i] = str(h)

    # Also grab attendance rate and truancy
    extra_cols = {}
    for i, h in enumerate(headers):
        if h and h in ('Student Attendance Rate', 'Chronic Truancy Rate'):
            extra_cols[i] = str(h)

    # All columns we want to read
    all_needed = set(id_cols.keys()) | set(ca_cols.keys()) | set(enrollment_cols.keys()) | set(extra_cols.keys())
    col_map = {**id_cols, **ca_cols, **enrollment_cols, **extra_cols}

    print(f"  Reading {len(all_needed)} columns from General sheet...")
    print(f"  Chronic Absenteeism columns found: {len(ca_cols)}")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {}
        for idx in all_needed:
            if idx < len(row):
                record[col_map[idx]] = row[idx]
        rows.append(record)

    wb.close()
    il = pd.DataFrame(rows)
    print(f"  Rows: {len(il):,}")

    # Save the wide-format Illinois data as-is for reference
    il['State'] = 'IL'
    il['Academic Year'] = '2023-24'
    return il


def pivot_illinois_to_long(il_wide):
    """
    Convert Illinois wide-format chronic absenteeism into long-format
    matching California's structure.

    Illinois columns like "Chronic Absenteeism - Male" become rows with
    Reporting Category = "Male".
    """
    # Mapping of IL column suffixes to a normalized reporting category
    ca_col_mapping = {
        'Chronic Absenteeism': 'TA',  # Total All
        'Chronic Absenteeism - Male': 'GM',
        'Chronic Absenteeism - Female': 'GF',
        'Chronic Absenteeism - White': 'RW',
        'Chronic Absenteeism - Black or African American': 'RB',
        'Chronic Absenteeism - Hispanic or Latino': 'RH',
        'Chronic Absenteeism - Asian': 'RA',
        'Chronic Absenteeism - Native Hawaiian or Other Pacific Islander': 'RP',
        'Chronic Absenteeism - American Indian or Alaska Native': 'RI',
        'Chronic Absenteeism - Two or More Races': 'RT',
        'Chronic Absenteeism - Children with Disabilities': 'SD',
        'Chronic Absenteeism - IEP': 'SIEP',
        'Chronic Absenteeism - EL': 'SEL',
        'Chronic Absenteeism - Low Income': 'SLI',
        'Chronic Absenteeism - Grade K': 'GRK',
        'Chronic Absenteeism - Grade 1': 'GR1',
        'Chronic Absenteeism - Grade 2': 'GR2',
        'Chronic Absenteeism - Grade 3': 'GR3',
        'Chronic Absenteeism - Grade 4': 'GR4',
        'Chronic Absenteeism - Grade 5': 'GR5',
        'Chronic Absenteeism - Grade 6': 'GR6',
        'Chronic Absenteeism - Grade 7': 'GR7',
        'Chronic Absenteeism - Grade 8': 'GR8',
        'Chronic Absenteeism - Grade 9': 'GR9',
        'Chronic Absenteeism - Grade 10': 'GR10',
        'Chronic Absenteeism - Grade 11': 'GR11',
        'Chronic Absenteeism - Grade 12': 'GR12',
    }

    id_columns = ['RCDTS', 'Type', 'School Name', 'District', 'City', 'County',
                  'District Type', 'School Type', 'Grades Served',
                  '# Student Enrollment', 'State', 'Academic Year',
                  'Student Attendance Rate', 'Chronic Truancy Rate']

    # Keep only columns that exist
    id_columns = [c for c in id_columns if c in il_wide.columns]

    records = []
    for _, row in il_wide.iterrows():
        base = {c: row.get(c) for c in id_columns}
        for il_col, reporting_cat in ca_col_mapping.items():
            if il_col in il_wide.columns:
                val = row.get(il_col)
                records.append({
                    **base,
                    'Reporting Category': reporting_cat,
                    'ChronicAbsenteeismRate': val
                })

    il_long = pd.DataFrame(records)

    # Clean rate values
    il_long['ChronicAbsenteeismRate'] = pd.to_numeric(
        il_long['ChronicAbsenteeismRate'].replace({'*': pd.NA, '': pd.NA}),
        errors='coerce'
    )

    print(f"  Illinois long-format rows: {len(il_long):,}")
    return il_long


# ---------------------------------------------------------------------------
# 3. Load into DuckDB
# ---------------------------------------------------------------------------
def load_into_duckdb(ca, il_wide, il_long):
    print(f"\nLoading into DuckDB at {DB_PATH}...")
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)

    con = duckdb.connect(DB_PATH)

    # California raw data
    con.execute("CREATE TABLE ca_chronic_absenteeism AS SELECT * FROM ca")
    count = con.execute("SELECT COUNT(*) FROM ca_chronic_absenteeism").fetchone()[0]
    print(f"  ca_chronic_absenteeism: {count:,} rows")

    # Illinois wide-format (for reference)
    con.execute("CREATE TABLE il_report_card_wide AS SELECT * FROM il_wide")
    count = con.execute("SELECT COUNT(*) FROM il_report_card_wide").fetchone()[0]
    print(f"  il_report_card_wide: {count:,} rows")

    # Illinois long-format (for comparison)
    con.execute("CREATE TABLE il_chronic_absenteeism_long AS SELECT * FROM il_long")
    count = con.execute("SELECT COUNT(*) FROM il_chronic_absenteeism_long").fetchone()[0]
    print(f"  il_chronic_absenteeism_long: {count:,} rows")

    # Also save CSVs for easy inspection
    ca.to_csv(os.path.join(DATA_DIR, 'ca_chronic_absenteeism_clean.csv'), index=False)
    il_long.to_csv(os.path.join(DATA_DIR, 'il_chronic_absenteeism_long.csv'), index=False)
    print("  Saved CSV exports.")

    con.close()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
if __name__ == '__main__':
    ca = load_california()
    il_wide = load_illinois()
    il_long = pivot_illinois_to_long(il_wide)
    load_into_duckdb(ca, il_wide, il_long)
    print("\nDone! Database ready at:", DB_PATH)
